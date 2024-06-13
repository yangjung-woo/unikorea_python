import openpyxl

def filter_military_terms(file_path, sheet_name, terms_to_filter):
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Find the last row in column A with data
    last_row = sheet.max_row
    
    # Iterate over each cell in column A up to the last row
    for row in range(1, last_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value:
            for term in terms_to_filter:
                if term in cell_value:
                    cell_value = cell_value.replace(term, 'OO')
            sheet.cell(row=row, column=1).value = cell_value
    
    # Save the changes to the Excel file
    workbook.save(file_path)
    print("군사 관련 용어 필터링 작업이 성공적으로 완료되었습니다!")

# Define file path and sheet name
file_path = 'C:/경로/스프레드시트.xlsx'  # 엑셀 파일 경로로 변경하세요
sheet_name = 'Sheet1'  # 워크시트 이름으로 변경하세요

# Define military-related terms to filter
terms_to_filter = ['군사', '전투기']  # 추가하고 싶은 용어가 있다면 여기에 추가하세요

# Run the function
filter_military_terms(file_path, sheet_name, terms_to_filter)
